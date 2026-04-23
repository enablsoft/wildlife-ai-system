from __future__ import annotations

"""
Frame export and species-label utility helpers.

Function index:
- last_taxon_segment: extract lowest-rank taxonomy segment.
- species_string_is_blank: detect blank species labels.
- record_is_blank: detect blank rows from species and description.
- format_species_display: format species label for UI text.
- short_species_label: build compact species label for tables.
- export_frames_xlsx: build Excel payload for videos/frames/species.
"""

from datetime import datetime, timezone
from io import BytesIO
from collections import Counter


def last_taxon_segment(species: str) -> str:
    """Return the last non-empty semicolon-delimited taxonomy segment."""
    for seg in reversed([p.strip() for p in (species or "").split(";")]):
        if seg:
            return seg
    return ""


def species_string_is_blank(species: str) -> bool:
    """Return True when species taxonomy indicates a blank/no-match record."""
    sp = (species or "").lower()
    if "__blank" in sp:
        return True
    return last_taxon_segment(species or "").lower() == "blank"


def record_is_blank(species: str, description: str) -> bool:
    """Return True when species or description marks a row as blank."""
    desc = (description or "").lower()
    if "__blank" in desc:
        return True
    return species_string_is_blank(species or "")


def format_species_display(species: str, description: str) -> str:
    """Format species text for UI display."""
    if record_is_blank(species, description):
        return "No species match (blank)"
    return short_species_label(species, description)


def short_species_label(species: str, description: str) -> str:
    """Build a compact species label for tables and exports."""
    if record_is_blank(species, description):
        return "Blank"
    seg = last_taxon_segment(species or "")
    if seg:
        return seg
    cleaned = (species or "").replace("_", " ").strip().title()
    return cleaned if cleaned and cleaned != "Unknown" else "Unknown"


def trailcam_stamp_label(row: dict[str, str]) -> str:
    """Build trail-cam stamp text from extracted overlay fields."""
    date_part = (row.get("overlay_date", "") or "").strip()
    time_part = (row.get("overlay_time", "") or "").strip()
    temp_part = (row.get("overlay_temp", "") or "").strip()
    parts = [p for p in (date_part, time_part, temp_part) if p]
    return " | ".join(parts)


def format_trailcam_temp(temp_raw: str) -> str:
    """Normalize OCR temperature text to plain numeric value."""
    temp = (temp_raw or "").strip()
    if not temp:
        return ""
    upper = temp.upper().replace("°", "")
    if upper.endswith("C") or upper.endswith("F"):
        upper = upper[:-1]
    return upper.strip()


def export_frames_xlsx(
    records: list[dict[str, str]],
    hide_blanks: bool,
    log_file: str,
) -> bytes:
    """Build an XLSX workbook for videos + frames + species and return bytes."""
    try:
        from openpyxl import Workbook  # type: ignore[reportMissingModuleSource]
    except Exception as e:  # pragma: no cover
        raise RuntimeError("openpyxl is required for Excel export.") from e
    wb = Workbook()
    frames_ws = wb.active
    frames_ws.title = "frames"
    frames_ws.append(
        [
            "video_source",
            "frame",
            "trail_cam_date",
            "trail_cam_time",
            "trail_cam_temp_(°C)",
            "species_label_short",
            "species_label_latin",
            "species_confidence",
            "default_species_short",
            "default_species_type",
            "species_taxonomy_full",
            "manual_tag",
            "description_species_context",
            "description_detector",
            "job_id",
        ]
    )
    for r in records:
        detector_class = r.get("detector_class", "")
        detector_conf = r.get("detector_confidence", "")
        detector_desc = f"{detector_class} ({detector_conf})".strip() if detector_conf else str(detector_class or "")
        frames_ws.append(
            [
                r.get("source", ""),
                r.get("frame", ""),
                r.get("overlay_date", ""),
                r.get("overlay_time", ""),
                format_trailcam_temp(r.get("overlay_temp", "")),
                short_species_label(r.get("species", ""), r.get("description", "")),
                r.get("species_latin", ""),
                r.get("species_confidence", ""),
                r.get("species_short", ""),
                r.get("species_type", ""),
                r.get("species", ""),
                r.get("manual_tag", ""),
                " ".join(
                    p
                    for p in [
                        f"Likely {short_species_label(r.get('species', ''), r.get('description', ''))}",
                        f"({r.get('species_latin', '')})" if r.get("species_latin", "") else "",
                        f"- confidence {r.get('species_confidence', '')}" if r.get("species_confidence", "") else "",
                    ]
                    if p
                ),
                detector_desc,
                r.get("job_id", ""),
            ]
        )
    frames_ws.freeze_panes = "A2"
    frames_ws.auto_filter.ref = f"A1:N{max(2, len(records) + 1)}"

    videos_ws = wb.create_sheet("videos")
    videos_ws.append(["video_source", "frame_count", "distinct_species_count", "blank_frame_count"])
    by_video: dict[str, list[dict[str, str]]] = {}
    for r in records:
        source = r.get("source", "") or ""
        by_video.setdefault(source, []).append(r)
    for source in sorted(by_video.keys(), key=lambda x: x.lower()):
        rows = by_video[source]
        species_set = {
            short_species_label(rr.get("species", ""), rr.get("description", ""))
            for rr in rows
            if not record_is_blank(rr.get("species", ""), rr.get("description", ""))
        }
        blank_count = sum(
            1 for rr in rows if record_is_blank(rr.get("species", ""), rr.get("description", ""))
        )
        videos_ws.append([source, len(rows), len(species_set), blank_count])
    videos_ws.freeze_panes = "A2"
    videos_ws.auto_filter.ref = f"A1:D{max(2, len(by_video) + 1)}"

    species_ws = wb.create_sheet("species")
    species_ws.append(["species_label_short", "frame_count"])
    species_counts: Counter[str] = Counter()
    for r in records:
        label = short_species_label(r.get("species", ""), r.get("description", ""))
        species_counts[label] += 1
    for label, count in sorted(species_counts.items(), key=lambda kv: (-kv[1], kv[0].lower())):
        species_ws.append([label, count])
    species_ws.freeze_panes = "A2"
    species_ws.auto_filter.ref = f"A1:B{max(2, len(species_counts) + 1)}"

    meta = wb.create_sheet("meta")
    meta.append(["generated_at_utc", datetime.now(timezone.utc).isoformat()])
    meta.append(["hide_blanks", "1" if hide_blanks else "0"])
    meta.append(["rows", str(len(records))])
    meta.append(["log_file", log_file])
    out = BytesIO()
    wb.save(out)
    return out.getvalue()

